import openpyxl

def abrir_excel():
    try:
        # Intentar abrir el archivo de Excel
        archivo = openpyxl.load_workbook('archivo.xlsx')
    except FileNotFoundError:
        return None

    # Buscar las hojas en el orden especificado
    for nombre_hoja in ['Base de datos', 'Base de datos ', 'DB']:
        if nombre_hoja in archivo.sheetnames:
            return archivo[nombre_hoja]

    # Si no se encontró ninguna hoja válida, devolver None
    return None

# Abrir el archivo Excel y obtener la hoja de trabajo
hoja_trabajo = abrir_excel()

if hoja_trabajo is not None:
    # Recorrer las filas
    for fila in hoja_trabajo.iter_rows(min_row=2, values_only=True):
        # Obtener el valor de la columna
        columna = fila[0]  # Suponiendo que la columna está en la posición 1
        
        print(columna)
        # Verificar si el valor es "Boleta"
        if columna == "Boleta":
            # Agregar el prefijo "Boleta "
            columna = "Boleta " + columna
        
        # Imprimir el valor actualizado
        print(columna)

    # Cerrar el archivo Excel
    hoja_trabajo.parent.close()
else:
    print("No se encontró una hoja válida en el archivo Excel.")
