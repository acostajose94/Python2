import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from funciones_columna import *
from datetime import datetime
import os
import shutil

 
archivo_excel = 'AVAL_TLS_PREGRADO NO  VIGENTE al  260523.xlsx'
datos_selec='datos_seleccionados NO VIGENTE al  260523.xlsx'
# Variables
cabecera = [
    'SOCIEDAD', 'EMPLID', 'CODIGO', 'APELLIDOS', 'NOMBRES', 'DIRECCION', 'DISTRITO', 'TELEFONO1', 'TELEFONO2',
    'TELEFONO3', 'TELEFONO4', 'EMAIL1', 'EMAIL2', 'EMAIL3', 'EMAIL4', 'PROGRAMA', 'CURSO', 'PERIODO',
    'NIVEL ACADEMICO', 'TIPO ALUMNO', 'DESCRIPCION', 'SALDO', 'MONEDA', 'ESTATUS CARRERA', 'MASIVA',
    'NUM. MASIVA', 'ITEM_NBR', 'DESCRIPCION ITEM', 'FECHA DEVENGO', 'FECHA EMISION', 'FECHA VENCIMIENTO',
    'NOTA CRED', 'ID FCRONOGRAMA', 'FECHA FIN PROGRAMA', 'CICLO', 'RAZON SOCIAL', 'RUC', 'TELEFONO EMPRESA',
    'FECHA CORTE', 'DESCR.', 'PALABRA CLAVE2', 'REF.', 'REF. FECHA', 'REF. COMENTARIOS', 'ID ORGANIZACION',
    'CANJE', 'SESION', 'TIPO PROGRAMA ', 'KOMODO', 'AGENCIA'
]
columnas_deseadas = ['CODIGO', 'TELEFONO3', 'TELEFONO4', 'EMAIL2', 'EMAIL3', 'EMAIL4', 'TIPO ALUMNO','NUM. MASIVA', 'DESCRIPCION']
nombres_nuevos_eliminar=['TELEFONO3','TELEFONO4','EMAIL2','EMAIL3','EMAIL4','NIVEL ACADEMICO','TIPO ALUMNO','ITEM_NBR','DESCRIPCION ITEM','FECHA DEVENGO','KOMODO']

# Abre el archivo de Excel
workbook = openpyxl.load_workbook(archivo_excel)
sheet = workbook['Hoja1']

verificar_cabecera_completa(sheet, cabecera)

datos_seleccionados = seleccionar_columnas_excel(sheet, columnas_deseadas)
# Crear un nuevo libro de Excel para almacenar los datos seleccionados
nuevo_libro = Workbook()
nueva_hoja = nuevo_libro.active

nueva_hoja=fusiona_agrega_BOLETA(nueva_hoja,columnas_deseadas,datos_seleccionados)        
# nueva_hoja=reorganize_sheet(nueva_hoja)
formato_texto(nueva_hoja,1,2,3)
# Guardar el nuevo archivo de Excel
nuevo_libro.save(datos_selec)

sheet = pasar_dato_desde_hasta(sheet,'NIVEL ACADEMICO', 'CICLO')
# Agregar columna en blanco en la tercera posición entre EMPLID y CODIGO
sheet.insert_cols(3)
sheet.insert_cols(7)
sheet.insert_cols(7)
sheet=eliminar_columnas(sheet,nombres_nuevos_eliminar)
formato_texto(sheet, 2, 4, 11, 12, 26, 30, 31, 38, 40)
formato_numero(sheet, 18)
formato_fecha(sheet, 23, 24, 27, 32)

# Cambiar la cabecera
cabecera_nueva = [
    'SOCIEDAD', 'EMPLID', 'UNICO', 'CODIGO', 'APELLIDOS', 'NOMBRES', 'TP_DOC_ID', 'DOC_ID',
    'DIRECCION', 'DISTRITO', 'TELEFONO', 'CELULAR', 'EMAIL', 'PROGRAMA', 'CURSO', 'PERIODO',
    'DESCRIPCION', 'SALDO', 'MONEDA', 'STATUS_CARRERA', 'MASIVA', 'NUM_MASIVA', 'FECHA_EMISION',
    'FECHA_VENCIMIENTO', 'NOTA_CRED', 'ID_FCRONOGRAMA', 'FECHA_FIN_PROGRANMA', 'CICLO',
    'RAZON_SOCIAL', 'RUC', 'PHONE_EMPRESA', 'FECHA_DE_CORTE', 'DESCR', 'PAL_CLAVE_2', 'REF',
    'REF_FECHA', 'REF_COMENTARIOS', 'ID_ORG_EXT', 'CANJE', 'SESION', 'TIPO_PROGRAMA', 'ASESOR'
]
 
# Asignar los valores de la nueva cabecera
for col_idx, valor in enumerate(cabecera_nueva, start=1):
    sheet.cell(row=1, column=col_idx, value=valor)

# Obtener la fecha actual en formato dd_mm
fecha_actual = datetime.now().strftime('%d_%m')

guardar_archivo_mover('B TLS', workbook, fecha_actual, 'B TLS VIGENTE',archivo_excel,datos_selec)
 