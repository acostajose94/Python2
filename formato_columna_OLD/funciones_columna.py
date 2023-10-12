from openpyxl.styles import numbers
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import shutil
def formato_texto(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]
    
    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            cell = row[0]
            cell.number_format = '@'
            if isinstance(cell.value, (int, float)):
                cell.value = str(cell.value)
            
def formato_numero(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]
        
    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            cell = row[0]
            valor = cell.value
            if valor is not None and isinstance(valor, str):
                valor = valor.replace(',', '.')
                cell.value = float(valor)
            cell.number_format = '0.00'

def formato_fecha(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]

    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            for cell in row:
                # Verificar si el valor termina en "/202"
                valor = str(cell.value)
                if valor.endswith("/202"):
                    nuevo_valor = valor[:-4] + "2"
                    cell.value = nuevo_valor

                # Aplicar formato de fecha
                cell.number_format = numbers.FORMAT_DATE_XLSX14

 
def guardar_archivo_mover(nombre_carpeta, workbook, fecha_actual, nombre_a_guardar, *archivos_mover):
    ruta_carpeta = os.path.join(os.path.join(os.environ['USERPROFILE']), 'OneDrive', 'Escritorio', fecha_actual + nombre_carpeta)
    if not os.path.exists(ruta_carpeta):
        os.makedirs(ruta_carpeta)

    ruta_archivo = os.path.join(ruta_carpeta, fecha_actual + nombre_a_guardar+'.xlsx')
    workbook.save(ruta_archivo)

    for arch in archivos_mover:
        shutil.move(arch, ruta_carpeta)
          
# Acepta listas NOMBRES
def eliminar_columnas(sheet, nombres_columnas):
    columnas_eliminar = []
    
    for columna in sheet.iter_cols():
        if columna[0].value in nombres_columnas:
            columnas_eliminar.append(columna)
    
    for columna in columnas_eliminar:
        sheet.delete_cols(columna[0].column)
    
    return sheet

# Acepta listas 
def seleccionar_columnas_excel(sheet, columnas_seleccionadas):
    columnas_seleccionadas_indices = []

    # Obtener el índice de cada columna seleccionada
    for columna in columnas_seleccionadas:
        for columna_actual in sheet.iter_cols(1, sheet.max_column):
            if columna_actual[0].value == columna:
                columnas_seleccionadas_indices.append(columna_actual[0].column)
                break

    # Crear una lista para almacenar los datos seleccionados
    datos_seleccionados = []

    # Recorrer las filas y obtener los datos de las columnas seleccionadas
    for fila in sheet.iter_rows(min_row=2, values_only=True):
        datos_fila = []
        for indice_columna in columnas_seleccionadas_indices:
            datos_fila.append(fila[indice_columna - 1])
        datos_seleccionados.append(datos_fila)

    return datos_seleccionados

def fusiona_agrega_BOLETA(nueva_hoja,columnas_deseadas,datos_seleccionados):
    # Escribir los encabezados de las columnas seleccionadas en el nuevo libro
    for indice, columna_seleccionada in enumerate(columnas_deseadas):
        letra_columna = get_column_letter(indice + 1)
        nueva_hoja[f'{letra_columna}1'] = str(columna_seleccionada)

    # Agregar una nueva columna llamada "BOLETA"
    letra_nueva_columna = get_column_letter(len(columnas_deseadas) + 1)
    nueva_hoja[f'{letra_nueva_columna}1'] = 'BOLETA'

    # Escribir los datos seleccionados y la columna "BOLETA" en el nuevo libro
    for indice_fila, fila_datos in enumerate(datos_seleccionados):
        for indice_columna, valor_celda in enumerate(fila_datos):
            letra_columna = get_column_letter(indice_columna + 1)
            nueva_hoja[f'{letra_columna}{indice_fila + 2}'] = valor_celda

        # Concatenar los valores de "NUM. MASIVA" y "DESCRIPCION" en la columna "BOLETA"
        num_masiva = fila_datos[7]  # Índice 7 corresponde a "NUM. MASIVA"
        descripcion = fila_datos[8]  # Índice 8 corresponde a "DESCRIPCION"
        boleta = f'{num_masiva}-{descripcion}'
        nueva_hoja[f'{letra_nueva_columna}{indice_fila + 2}'] = boleta
        
    return nueva_hoja

# def reorganize_sheet(sheet):
#     # Obtener las filas del archivo
#     rows = sheet.iter_rows(values_only=True)

#     # Obtener los encabezados de la primera fila
#     headers = next(rows)

#     # Obtener los índices de las columnas a modificar
#     codigo_idx = headers.index('CODIGO')
#     telefono3_idx = headers.index('TELEFONO3')
#     telefono4_idx = headers.index('TELEFONO4')
#     email2_idx = headers.index('EMAIL2')
#     email3_idx = headers.index('EMAIL3')
#     email4_idx = headers.index('EMAIL4')

#     # Crear una nueva lista de encabezados modificados
#     new_headers = ['CODIGO', 'TELEFONO3', 'EMAIL2', 'TIPO ALUMNO', 'NUM. MASIVA', 'DESCRIPCION', 'BOLETA']

#     # Crear una nueva lista para almacenar las filas modificadas
#     new_rows = [new_headers]

#     # Recorrer las filas restantes y realizar los cambios necesarios
#     for row in rows:
#         codigo = row[codigo_idx]
#         telefono3 = row[telefono3_idx]
#         telefono4 = row[telefono4_idx]
#         email2 = row[email2_idx]
#         email3 = row[email3_idx]
#         email4 = row[email4_idx]

#         new_row = [codigo, telefono3, email2, '', '', '', '']
#         new_rows.append(new_row)

#         if telefono4:
#             new_row = [codigo, telefono4, email3, '', '', '', '']
#             new_rows.append(new_row)

#         if email4:
#             new_row = [codigo, telefono4, email4, '', '', '', '']
#             new_rows.append(new_row)

#     # Limpiar el contenido actual de la hoja
#     sheet.delete_rows(1, sheet.max_row)

#     # Agregar las filas modificadas a la hoja
#     for new_row in new_rows:
#         sheet.append(new_row)

#     return sheet

def pasar_dato_desde_hasta(sheet, dato_desde, dato_para):
    columna_nivel_academico = None
    columna_ciclo = None

    # Buscar las columnas por su nombre
    for columna in sheet.iter_cols():
        if columna[0].value == dato_desde:
            columna_nivel_academico = columna[0].column_letter
        elif columna[0].value == dato_para:
            columna_ciclo = columna[0].column_letter

    # Verificar que se encontraron las columnas
    if columna_nivel_academico and columna_ciclo:
        for fila in range(2, sheet.max_row + 1):
            nivel_academico = sheet[columna_nivel_academico + str(fila)].value
            if nivel_academico is not None:
                ciclo = nivel_academico.split()[0]
                sheet[columna_ciclo + str(fila)].value = ciclo
            else:
                sheet[columna_ciclo + str(fila)].value = None

        return sheet
    else:
        print("No se encontraron las columnas especificadas.")
        return None

def verificar_cabecera_completa(sheet, cabecera):
    # Obtener la primera fila de la hoja de trabajo (cabecera)
    primera_fila = sheet[1]
    
    # Verificar si la cabecera completa existe
    for celda, valor_cabecera in zip(primera_fila, cabecera):
        if celda.value != valor_cabecera:
            print(f"La columna {celda.column_letter} no coincide: se esperaba '{valor_cabecera}', se encontró '{celda.value}'")
            return False
            break
    
    return True

