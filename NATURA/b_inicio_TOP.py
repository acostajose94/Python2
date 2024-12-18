import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from b_funciones_NAT import *
from b_variables_TOP import *

# archivo_reciente=extraer_archivo_mas_reciente()
archivos_excel=listar_excels()

ruta_carpeta=define_carpeta('NAT TOP',fecha_actual)

for archivo_excel in archivos_excel:

    pre_workbook=load_workbook(filename=archivo_excel)
    sheet_PRE = pre_workbook.active
    verificar_cabecera_completa(sheet_PRE, cab_esperada)
 
    workbook=reordenar_columnas(pre_workbook,cab_ordenada)
    sheet = workbook.active
    
    formato_texto(sheet,1,2,3,5,6,7,22,25,27,29, 31, 32)
    formato_numero(sheet, 11,12,13,14,16,17,18,19,20)
    formato_fecha(sheet, 8,9,10,21,24)

    # Asignar los valores de la nueva cabecera
    for col_idx, valor in enumerate(cab_nueva, start=1):
        sheet.cell(row=1, column=col_idx, value=valor)

    columna_letra = 'AD'
    # Recorrer la columna y eliminar los caracteres especificados
    for celda in sheet[columna_letra]:
        # Reemplazar los caracteres por una cadena vacía
        celda.value = str(celda.value).replace(chr(10), '').replace(chr(9), '').replace(chr(13), '')


    ruta_base=os.path.join(ruta_carpeta, archivo_excel)
    workbook.save(ruta_base + '.xlsx')
    shutil.move(archivo_excel,ruta_carpeta)

    print(ruta_base)
 
print('Echo')
 