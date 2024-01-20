import openpyxl

# Cargar el archivo Excel
archivo = "20231212. SALDO ACTUALIZADO 1-1800 DIC - AVAL.xlsx"
wb = openpyxl.load_workbook(archivo)
sheet = wb.active

# Agregar columna 'NO_HAY' en el encabezado
nueva_columna = 'NO_HAY'
sheet.insert_cols(5)  # Insertar en la posición deseada
sheet.cell(row=1, column=5).value = nueva_columna

# Definir el nuevo orden de las columnas
nuevo_orden = [
    'Orden del Pedido',
    'Código Consultora',
    'Grupo Lider',
    'Nombre Consultora',
    
    'Nro Pedido',
    
    'Fecha Pedido',
    'Fecha Vcto',
    
    'Valor Original',
    'Saldo',
    'Saldo Corregido',
    'Interés/Multa Acumulados',
    '% Descuento Saldo',
    'Monto Descuento Saldo',
    'Total Descuentos',
    'Monto mínimo a pagar Campaña',
    
    'Días de Retraso',
    'Ciclo',
    'Fecha Nacimiento',
    
    
    
    
    'DNI',
    'Dirección',
    'Teléfono 1',
    'Teléfono 2',
    'Correo',
    'Tramo Vencimiento',
    'Agencia',
    'Riesgo'
]

# Crear un nuevo libro y una nueva hoja
nuevo_libro = openpyxl.Workbook()
nueva_hoja = nuevo_libro.active

# Copiar los datos reordenados
for col_idx, header in enumerate(nuevo_orden, start=1):
    columna_origen = sheet[openpyxl.utils.get_column_letter(sheet[header].column)]
    for row_idx, cell in enumerate(columna_origen, start=1):
        nueva_celda = nueva_hoja.cell(row=row_idx, column=col_idx)
        nueva_celda.value = cell.value

# Guardar los cambios en un nuevo archivo
nuevo_libro.save("20231212.SALDO_ACTUALIZADO_reorganizado.xlsx")
