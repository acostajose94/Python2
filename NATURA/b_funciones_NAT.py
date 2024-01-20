from openpyxl.styles import numbers
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
import os
import shutil
import glob
import zipfile
import rarfile

def formato_texto(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]
    
    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            cell = row[0]
            cell.number_format = '@'
            if isinstance(cell.value, (int, float)):
                cell.value = str(cell.value)
            
def formato_numero(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]
        
    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            cell = row[0]
            valor = cell.value
            if valor is not None and isinstance(valor, str):
                valor = valor.replace(',', '.')
                valor = valor.replace('-', '0')
                cell.value = float(valor)
            cell.number_format = '0.00'

def formato_fecha(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]

    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            for cell in row:
                # Verificar si el valor termina en "/202"
                valor = str(cell.value)
                if valor.endswith("/202"):
                    nuevo_valor = valor[:-4] + "2"
                    cell.value = nuevo_valor

                # Aplicar formato de fecha
                cell.number_format = numbers.FORMAT_DATE_XLSX14


def extraer_archivo_mas_reciente():
    # Obtener la ruta del archivo Python actual
    ruta_actual = os.path.dirname(os.path.abspath(__file__))

    # Buscar archivos ZIP y RAR en la carpeta actual
    archivos_zip = glob.glob(os.path.join(ruta_actual, '*.zip'))
    archivos_rar = glob.glob(os.path.join(ruta_actual, '*.rar'))

    # Combinar y ordenar los archivos por fecha de modificación descendente
    archivos = sorted(archivos_zip + archivos_rar, key=os.path.getmtime, reverse=True)

    if archivos:
        archivo_mas_reciente = archivos[0]
        nombre_archivo, extension = os.path.splitext(archivo_mas_reciente)

        # Extraer archivo ZIP
        if extension == '.zip':
            with zipfile.ZipFile(archivo_mas_reciente, 'r') as archivo_zip:
                archivo_zip.extractall(ruta_actual)
                print(f'Se ha extraído el archivo ZIP más reciente: {archivo_mas_reciente}')

        # Extraer archivo RAR
        elif extension == '.rar':
            with rarfile.RarFile(archivo_mas_reciente, 'r') as archivo_rar:
                archivo_rar.extractall(ruta_actual)
                print(f'Se ha extraído el archivo RAR más reciente: {archivo_mas_reciente}')

        return archivo_mas_reciente

    else:
        print('No se encontraron archivos ZIP o RAR en la carpeta actual.')
   
def listar_excels():
    archivos_excel = []
    directorio_actual = os.getcwd()

    for archivo in os.listdir(directorio_actual):
        if archivo.endswith('.xlsx') or archivo.endswith('.xls'):
            if 'Retira' not in archivo:
                archivos_excel.append(archivo)
    
    return archivos_excel     

 
def define_carpeta(nombre_carpeta,fecha_actual):
    ruta_carpeta = os.path.join(os.path.join(os.environ['USERPROFILE']), 'OneDrive', 'Escritorio', fecha_actual + nombre_carpeta)
    if not os.path.exists(ruta_carpeta):
        os.makedirs(ruta_carpeta)
    return ruta_carpeta

 
def guardar_archivo_mover(nombre_carpeta, workbook, fecha_actual, nombre_a_guardar, *archivos_mover):
    ruta_carpeta = os.path.join(os.path.join(os.environ['USERPROFILE']), 'OneDrive', 'Escritorio', fecha_actual + nombre_carpeta)
    if not os.path.exists(ruta_carpeta):
        os.makedirs(ruta_carpeta)

    ruta_archivo = os.path.join(ruta_carpeta, 'BASE' + nombre_a_guardar)
    workbook.save(ruta_archivo)

    for arch in archivos_mover:
        shutil.move(arch, ruta_carpeta)
   
def reordenar_columnas(archivo_excel, nuevo_orden):
    # Obtener la hoja pasada como argumento
    libro = load_workbook(filename=archivo_excel)
    hoja = libro.active

    # Insertar una columna 'NO_HAY_DATA' al principio sin datos
    hoja.insert_cols(1)
    hoja.cell(row=1, column=1, value='NO_HAY_DATA')

    # Obtener las posiciones actuales de las cabeceras
    cabeceras = [hoja.cell(row=1, column=col).value for col in range(1, hoja.max_column + 1)]
    posiciones_actuales = {cabecera: index + 1 for index, cabecera in enumerate(cabeceras)}

    # Crear una nueva hoja para almacenar los datos reordenados
    nuevo_libro = Workbook()
    nueva_hoja = nuevo_libro.active

    # Crear las cabeceras en el nuevo orden
    for cabecera in nuevo_orden:
        nueva_hoja.cell(row=1, column=nuevo_orden.index(cabecera) + 1, value=cabecera)

    # Copiar los datos a la nueva hoja en el nuevo orden
    for row in hoja.iter_rows(min_row=2, max_row=hoja.max_row, values_only=True):
        nueva_fila = []
        for cabecera in nuevo_orden:
            nueva_fila.append(row[posiciones_actuales[cabecera] - 1])
        nueva_hoja.append(nueva_fila)

    return nuevo_libro   
       
# Acepta listas NOMBRES
def eliminar_columnas(sheet, nombres_columnas):
    columnas_eliminar = []
    
    for columna in sheet.iter_cols():
        if columna[0].value in nombres_columnas:
            columnas_eliminar.append(columna)
    
    for columna in columnas_eliminar:
        sheet.delete_cols(columna[0].column)
    
    return sheet

# Acepta listas , para obtener datos de las columnas
def seleccionar_columnas_excel(sheet, columnas_seleccionadas):
    columnas_seleccionadas_indices = []

    # Obtener el índice de cada columna seleccionada
    for columna in columnas_seleccionadas:
        for columna_actual in sheet.iter_cols(1, sheet.max_column):
            if columna_actual[0].value == columna:
                columnas_seleccionadas_indices.append(columna_actual[0].column)
                break

    # Crear una lista para almacenar los datos seleccionados
    datos_seleccionados = []

    # Recorrer las filas y obtener los datos de las columnas seleccionadas
    for fila in sheet.iter_rows(min_row=2, values_only=True):
        datos_fila = []
        for indice_columna in columnas_seleccionadas_indices:
            datos_fila.append(fila[indice_columna - 1])
        datos_seleccionados.append(datos_fila)

    return datos_seleccionados
 
def pasar_dato_desde_hasta(sheet, dato_desde, dato_para):
    columna_nivel_academico = None
    columna_ciclo = None

    # Buscar las columnas por su nombre
    for columna in sheet.iter_cols():
        if columna[0].value == dato_desde:
            columna_nivel_academico = columna[0].column_letter
        elif columna[0].value == dato_para:
            columna_ciclo = columna[0].column_letter

    # Verificar que se encontraron las columnas
    if columna_nivel_academico and columna_ciclo:
        for fila in range(2, sheet.max_row + 1):
            nivel_academico = sheet[columna_nivel_academico + str(fila)].value
            if nivel_academico is not None:
                ciclo = nivel_academico.split()[0]
                sheet[columna_ciclo + str(fila)].value = ciclo
            else:
                sheet[columna_ciclo + str(fila)].value = None

        return sheet
    else:
        print("No se encontraron las columnas especificadas.")
        return None

def verificar_cabecera_completa(sheet, cabecera):
    # Obtener la primera fila de la hoja de trabajo (cabecera)
    primera_fila = sheet[1]
    
    # Verificar si la cabecera completa existe
    for celda, valor_cabecera in zip(primera_fila, cabecera):
        if celda.value != valor_cabecera:
            print(f"La columna {celda.column_letter} no coincide: se esperaba '{valor_cabecera}', se encontró '{celda.value}'")
            return False
            break
    
    return True


def buscar_posicion_por_nombre(sheet, nombre_columna):
    for columna in sheet.iter_cols(min_row=1, max_row=1):
        for celda in columna:
            if celda.value == nombre_columna:
                return celda.column  # Devuelve el número de columna
    return None  # Si no se encuentra, devuelve None

# Define una función para eliminar columnas por nombre
def eliminar_columnas_por_nombre(sheet, columnas_a_eliminar):
    for col_name in columnas_a_eliminar:
        for col in sheet.columns:
            if col[0].value == col_name:
                sheet.delete_cols(col[0].column)

def elimina_tabulaciones(sheet: Worksheet, *columnas):
    for col_idx in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell,Cell) and cell.value is not None and '\t\n' in cell.value:
                    cell.value = str(cell.value).replace(chr(10), '').replace(chr(9), '').replace(chr(13), '')

def eliminar_espacios_y_guiones(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]

    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            for cell in row:
                if isinstance(cell.value, str):
                    # Reemplazar guión medio y espacios en blanco con cadena vacía
                    cell.value = cell.value.replace('-', '').replace(' ', '')
