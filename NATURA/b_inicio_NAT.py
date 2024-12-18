import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from b_funciones_NAT import *
from b_variables_NAT import *

def mover_contenido_largo_por_letras(sheet, col_telefono, col_celular, col_direccion):
    for row in range(2, sheet.max_row + 1):  # Empezamos desde la fila 2 para evitar la cabecera
        telefono_valor = sheet.cell(row=row, column=col_telefono).value
        celular_valor = sheet.cell(row=row, column=col_celular).value

        def debe_moverse(valor):
            if valor:
                valor_str = str(valor)
                letras = sum(c.isalpha() for c in valor_str)
                numeros = sum(c.isdigit() for c in valor_str)
                return letras > 10 and numeros < 5
            return False

        if debe_moverse(telefono_valor):
            direccion_actual = sheet.cell(row=row, column=col_direccion).value
            nueva_direccion = (direccion_actual or '') + ' ' + str(telefono_valor)
            sheet.cell(row=row, column=col_direccion).value = nueva_direccion.strip()
            sheet.cell(row=row, column=col_telefono).value = None  # Borrar el contenido de TELEFONO_PERSONA

        if debe_moverse(celular_valor):
            direccion_actual = sheet.cell(row=row, column=col_direccion).value
            nueva_direccion = (direccion_actual or '') + ' ' + str(celular_valor)
            sheet.cell(row=row, column=col_direccion).value = nueva_direccion.strip()
            sheet.cell(row=row, column=col_celular).value = None  # Borrar el contenido de NUMERO_CELULAR



# archivo_reciente=extraer_archivo_mas_reciente()
archivos_excel=listar_excels()

ruta_carpeta=define_carpeta('NAT S',fecha_actual)

for archivo_excel in archivos_excel:

    # pre_workbook=load_workbook(filename=archivo_excel)
    # workbook=reordenar_columnas(pre_workbook,nuevo_orden)
    workbook_PREV=load_workbook(filename=archivo_excel)
    sheet = workbook_PREV.active
    verificar_cabecera_completa(sheet, cab_esperada)
 
    
    workbook=reordenar_columnas(workbook_PREV,cab_ordenada)
    sheet = workbook.active
    # # Agregar columna en blanco en la tercera posición entre EMPLID y CODIGO
    # sheet.insert_cols(19)
    # sheet.insert_cols(19)
    
    formato_texto(sheet,1,2,3,5,6,7,22,25,27,29, 31, 32)
    formato_numero(sheet, 11,12,13,14,16,17,18,19,20)
    formato_fecha(sheet, 8,9,10,21,24)

    # Asignar los valores de la nueva cabecera
    for col_idx, valor in enumerate(cabecera_nueva, start=1):
        sheet.cell(row=1, column=col_idx, value=valor)

    columna_letra = 'AD'
    # Recorrer la columna y eliminar los caracteres especificados
    for celda in sheet[columna_letra]:
        # Reemplazar los caracteres por una cadena vacía
        celda.value = str(celda.value).replace(chr(10), '').replace(chr(9), '').replace(chr(13), '')

    # Buscar posiciones de columnas por nombre
    col_telefono = buscar_posicion_por_nombre(sheet, 'TELEFONO_PERSONA')
    col_celular = buscar_posicion_por_nombre(sheet, 'NUMERO_CELULAR')
    col_direccion = buscar_posicion_por_nombre(sheet, 'DIRECCION')

    if col_telefono and col_celular and col_direccion:
        mover_contenido_largo_por_letras(sheet, col_telefono, col_celular, col_direccion)



    ruta_base=os.path.join(ruta_carpeta, archivo_excel)
    workbook.save(ruta_base + '.xlsx')
    shutil.move(archivo_excel,ruta_carpeta)

    print(ruta_base)
 
print('Echo')
 