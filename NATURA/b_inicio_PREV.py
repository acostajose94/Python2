import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from b_funciones_NAT import *
from b_variables_PREV import *

# archivo_reciente=extraer_archivo_mas_reciente()
archivos_excel=listar_excels()

ruta_carpeta=define_carpeta('NAT PREV',fecha_actual)

for archivo_excel in archivos_excel:
    #-------EXCEL PARA COMENTARIO---------#
    #Nuebo book para coment 
    nuevo_wb =  load_workbook(filename=archivo_excel)
    sheet_PRE = nuevo_wb.active
    verificar_cabecera_completa(sheet_PRE, cab_esperada)
    
    nuevo_wb=reordenar_columnas(nuevo_wb,orden_cabecera_comentario)
    com_sheet = nuevo_wb.active
    com_sheet.title = "Hoja1"
    
    formato_texto(com_sheet,1,2)
    
    for col_idx, valor in enumerate(cabecera_comentario, start=1):
        com_sheet.cell(row=1, column=col_idx, value=valor)
    ruta_base_com=os.path.join(ruta_carpeta, 'COMENT')
    nuevo_wb.save(ruta_base_com + '.xlsx')
    #-------EXCEL PARA PAGO---------#
    #Nuebo book para PAGO
    nuevo_wb2 =   load_workbook(filename=archivo_excel)
    nuevo_wb2=reordenar_columnas(nuevo_wb2,orden_cabecera_pagos)
    nuevo_sheet2 = nuevo_wb2.active
     
    nuevo_sheet2.title = "Hoja1"
    
    
    rellenar(nuevo_sheet2, fecha_dmy, 8)
    rellenar(nuevo_sheet2, 'AVAL', 10)
    
    formato_texto(nuevo_sheet2,1,6)
    formato_numero(nuevo_sheet2, 4)
    formato_fecha(nuevo_sheet2,3,8)
    
    
    for col_idx, valor in enumerate(cabecera_pagos, start=1):
        nuevo_sheet2.cell(row=1, column=col_idx, value=valor)
    
    ruta_base_com2=os.path.join(ruta_carpeta, 'PAGO')
    nuevo_wb2.save(ruta_base_com2 + '.xlsx')
    
    
   
    # # Ajuste  de saldos
    def corregir_monto(workbook_path):
        wb = load_workbook(filename=workbook_path)
        sheet = wb.active
        
        for row in sheet.iter_rows(min_row=2):
            saldo_actualizado_vf = row[21].value
            
            if saldo_actualizado_vf is None:
                monto_recuperado = row[19].value
                pago = row[16].value
                saldo_actualizado = row[6].value
                
                
                if pago is None:
                    pago=0
                saldo_actualizado_vf = saldo_actualizado - pago
                if saldo_actualizado_vf < 0:
                    saldo_actualizado_vf=0
                row[21].value = saldo_actualizado_vf
            
            # elif saldo_actualizado_vf=="-":
            #     saldo_actualizado = row[6].value
            #     saldo_actualizado_vf=saldo_actualizado
            elif saldo_actualizado_vf <0:
                saldo_actualizado_vf=0
                row[21].value = saldo_actualizado_vf
        return wb

    pre_workbook=corregir_monto(archivo_excel)
    pre_sheet=pre_workbook.active
    
    
    
    
    
    # Reordena excel
    workbook=reordenar_columnas(pre_workbook,nuevo_orden_prev)
    sheet = workbook.active
    
    
    
    
    rellenar(sheet, 0, 14,15,16,17)
    rellenar(sheet, 'AVAL', 35)

    # aplicar_reemplazo(sheet, '', '0', 0,1,2,3,4,5)
    
    
    formato_texto(sheet,1,2,3,5,6,7,22,25,27,29, 31, 32)
    formato_numero(sheet, 11,12,13,14,16,17,18,19,20)
    formato_fecha(sheet, 8,9,10,21,24)

    # Asignar los valores de la nueva cabecera
    for col_idx, valor in enumerate(cabecera_nueva, start=1):
        sheet.cell(row=1, column=col_idx, value=valor)

    # columna_letra = 'AD'
    # # Recorrer la columna y eliminar los caracteres especificados
    # for celda in sheet[columna_letra]:
    #     # Reemplazar los caracteres por una cadena vacía
    #     celda.value = str(celda.value).replace(chr(10), '').replace(chr(9), '').replace(chr(13), '')


    ruta_base=os.path.join(ruta_carpeta, archivo_excel)
    workbook.save(ruta_base + '.xlsx')
    shutil.move(archivo_excel,ruta_carpeta)

    print(ruta_base)
 
print('Echo')
 