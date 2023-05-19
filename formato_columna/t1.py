import openpyxl
from openpyxl.styles import numbers

# Reemplaza 'archivo.xlsx' con el nombre de tu archivo de Excel
archivo_excel = 'archivo.xlsx'

# Abre el archivo de Excel
workbook = openpyxl.load_workbook('Libro1.xlsx')
sheet = workbook.active

# Establece el formato de la columna B como texto
for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
    cell = row[0]
    cell.number_format = '@'

# Establece el formato de la columna C como fecha corta
for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
    cell = row[0]
    cell.number_format = numbers.FORMAT_DATE_XLSX14

# Establece el formato de la columna D como número
for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
    cell = row[0]
    cell.number_format = '0.00'

# Guarda los cambios en el archivo de Excel
workbook.save(archivo_excel)
