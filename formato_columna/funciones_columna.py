from openpyxl.styles import numbers

def formato_texto(sheet, columna):
    # for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
    #     cell = row[0]
    #     cell.number_format = '@'
    
    # for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
    #     cell = row[0]
    #     cell.number_format = '@'
    #     if isinstance(cell.value, (int, float)):
    #         cell.value = str(cell.value)
    #         cell.number_format = 'General'
    for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
        cell = row[0]
        cell.number_format = '@'
        if isinstance(cell.value, (int, float)):
            cell.value = str(cell.value)       
            
def formato_fecha(sheet, columna):
    for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
        cell = row[0]
        cell.number_format = numbers.FORMAT_DATE_XLSX14

def formato_numero(sheet, columna):
    for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
        cell = row[0]
        valor = cell.value
        if valor is not None and isinstance(valor, str):
            valor = valor.replace(',', '.')
            cell.value = float(valor)
        cell.number_format = '0.00'
