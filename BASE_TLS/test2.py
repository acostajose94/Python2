import openpyxl

# Cargar el archivo de Excel
archivo_excel = openpyxl.load_workbook('SELEC VIGENTE  _06.06.23.xlsx')
# Seleccionar la hoja de trabajo
hoja_trabajo = archivo_excel.active

# Crear una nueva hoja de trabajo para el resultado
hoja_resultado = archivo_excel.create_sheet("Resultado")

# Obtener los nombres de las columnas
columnas = hoja_trabajo[1]
nombres_columnas = [columna.value for columna in columnas]

# Obtener los índices de las columnas requeridas
indice_emplid = nombres_columnas.index('EMPLID')
indice_telefono3 = nombres_columnas.index('TELEFONO3')
indice_telefono4 = nombres_columnas.index('TELEFONO4')

# Escribir los encabezados en la hoja de resultado
hoja_resultado['A1'] = "EMPLID"
hoja_resultado['B1'] = "TELEFONOS"

# Variable para realizar el seguimiento de la última fila escrita para cada emplid
ultima_fila_emplid = {}

# Lista para almacenar las filas no vacías
filas_no_vacias = []

# Recorrer las filas de la hoja de trabajo original
for fila in hoja_trabajo.iter_rows(min_row=2, values_only=True):
    emplid = fila[indice_emplid]
    telefono3 = fila[indice_telefono3]
    telefono4 = fila[indice_telefono4]

    if telefono3:
        filas_no_vacias.append((emplid, telefono3))
    
    if telefono4:
        filas_no_vacias.append((emplid, telefono4))

# Escribir las filas no vacías en la hoja de resultado
for i, fila in enumerate(filas_no_vacias, start=2):
    hoja_resultado.cell(row=i, column=1).value = fila[0]
    hoja_resultado.cell(row=i, column=2).value = fila[1]

# Guardar el archivo con el nuevo formato
archivo_excel.save('archivo_open.xlsx')