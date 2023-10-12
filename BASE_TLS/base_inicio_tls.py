import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from base_funciones_tls import *
from base_variables_tls import *

archivo_reciente=extraer_archivo_mas_reciente()
archivos_excel=listar_excels()
retira=obtener_excel_retira()
shutil.copy(retira, destino_retirar)
ruta_carpeta=define_carpeta('B TLS',fecha_actual)

for archivo_excel in archivos_excel:
    # Abre el archivo de Excel
    workbook = openpyxl.load_workbook(archivo_excel)
    sheet = workbook['Hoja1']

    verificar_cabecera_completa(sheet, cabecera)
    # datos_seleccionados = seleccionar_columnas_excel(sheet, columnas_deseadas)
    # nuevo_libro = Workbook()
    # nueva_hoja = nuevo_libro.active

    # nueva_hoja=fusiona_agrega_BOLETA(nueva_hoja,columnas_deseadas,datos_seleccionados)        
    # formato_texto(nueva_hoja,1,2,3)
    # indice_inicio = archivo_excel[archivo_excel.find('AVAL_TLS_PREGRADO') + len('AVAL_TLS_PREGRADO'):]
    # selec='SELEC'+indice_inicio
    # nuevo_libro.save(selec)


    columnas_a_eliminar = ['COD. DUPLI','SEGMENTACIÓN']
    eliminar_columnas_por_nombre(sheet, columnas_a_eliminar)

    sheet = pasar_dato_desde_hasta(sheet,'NIVEL ACADEMICO', 'CICLO')
    # Agregar columna en blanco en la tercera posición entre EMPLID y CODIGO
    sheet.insert_cols(3)
    sheet.insert_cols(7)
    sheet.insert_cols(7)
   
    formato_texto(sheet, 2, 4, 11, 12,13,14,30, 36, 38, 40, 41,48,50)
    formato_numero(sheet, 25)
    formato_fecha(sheet, 32,33,34,37,42)

    # Asignar los valores de la nueva cabecera
    for col_idx, valor in enumerate(cabecera_nueva, start=1):
        sheet.cell(row=1, column=col_idx, value=valor)

    nombre_short=obtener_despues_de_pregrado(archivo_excel)
    ruta_base=os.path.join(ruta_carpeta, 'BASE TLS ' + nombre_short)
    workbook.save(ruta_base + '.xlsx')
    shutil.move(archivo_excel,ruta_carpeta)

    print(nombre_short)

    # print('Guardando')
    # guardar_archivo_mover('B TLS', workbook, hoy_carpeta,indice_inicio,archivo_excel,selec)
shutil.move(retira,ruta_carpeta)
print('Echo')

os.remove(archivo_reciente)