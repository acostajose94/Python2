import glob
import os
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.worksheet.worksheet import Worksheet
import shutil
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

import os
import zipfile
import rarfile
from pagos_variables import *


def eliminar_archivos_excel(ruta_actual):
     
    archivos = os.listdir(ruta_actual)
    
    for archivo in archivos:
        if archivo.endswith(".xlsx") or archivo.endswith(".xls"):
            archivo_path = os.path.join(ruta_actual, archivo)
            os.remove(archivo_path)
            print(f"Archivo eliminado: {archivo_path}")

def extraer_archivo_mas_reciente(ruta_actual):
 

    # Buscar archivos ZIP y RAR en la carpeta actual
    archivos_zip = glob.glob(os.path.join(ruta_actual, '*.zip'))
    archivos_rar = glob.glob(os.path.join(ruta_actual, '*.rar'))

    # Combinar y ordenar los archivos por fecha de modificación descendente
    archivos = sorted(archivos_zip + archivos_rar, key=os.path.getmtime, reverse=True)

    if archivos:
        archivo_mas_reciente = archivos[0]
        nombre_archivo, extension = os.path.splitext(archivo_mas_reciente)

        # Extraer archivo ZIP
        if extension == '.zip':
            with zipfile.ZipFile(archivo_mas_reciente, 'r') as archivo_zip:
                archivo_zip.extractall(ruta_actual)
                print(f'Se ha extraído el archivo ZIP más reciente: {archivo_mas_reciente}')

        # Extraer archivo RAR
        elif extension == '.rar':
            with rarfile.RarFile(archivo_mas_reciente, 'r') as archivo_rar:
                archivo_rar.extractall(ruta_actual)
                print(f'Se ha extraído el archivo RAR más reciente: {archivo_mas_reciente}')
        
        return archivo_mas_reciente

    else:
        print('No se encontraron archivos ZIP o RAR en la carpeta actual.')
   

def buscar_archivo_mas_reciente(directorio, palabra):
    # Buscar archivos que contengan la palabra
    archivos = glob.glob(os.path.join(directorio, f'*{palabra}*'))
    
    # Verificar si se encontraron archivos
    if not archivos:
        return None
    
    # Obtener el archivo más reciente
    archivo_mas_reciente = max(archivos, key=os.path.getmtime)
    
    return archivo_mas_reciente

def formato_texto(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]
    
    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            cell = row[0]
            cell.number_format = '@'
            if isinstance(cell.value, (int, float)):
                cell.value = str(cell.value)
            
def formato_numero(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]
        
    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            cell = row[0]
            valor = cell.value
            if valor is not None and isinstance(valor, str):
                valor = valor.replace(',', '.')
                cell.value = float(valor)
            cell.number_format = '0.00'

def formato_fecha(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]

    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            for cell in row:
                # Verificar si el valor termina en "/202"
                valor = str(cell.value)
                if valor.endswith("/202"):
                    nuevo_valor = valor[:-4] + "2"
                    cell.value = nuevo_valor

                # Aplicar formato de fecha
                cell.number_format = numbers.FORMAT_DATE_XLSX14

def verificar_cabecera(sheet, cabecera_esperada):
    header_row = sheet[1]

    for cell, expected_value in zip(header_row, cabecera_esperada):
        if cell.value != expected_value:
            print(f"La cabecera esperada no coincide en la columna {cell.column}. Se esperaba '{expected_value}', pero se encontró '{cell.value}'.")
            return False

    return True

def define_carpeta(nombre_carpeta,fecha_actual):
    ruta_carpeta = os.path.join(os.path.join(os.environ['USERPROFILE']), 'OneDrive', 'Escritorio', fecha_actual + nombre_carpeta)
    if not os.path.exists(ruta_carpeta):
        os.makedirs(ruta_carpeta)
    return ruta_carpeta

def buscar_posicion_por_nombre(sheet, nombre_columna):
    for columna in sheet.iter_cols(min_row=1, max_row=1):
        for celda in columna:
            if celda.value == nombre_columna:
                return celda.column  # Devuelve el número de columna
    return None  # Si no se encuentra, devuelve None


def pagos_certus(filename, cabecera_esperada, nueva_cabecera,fecha_actual):
    wb = load_workbook(filename)
    sheet = wb.active

    if verificar_cabecera(sheet, cabecera_esperada):
        print("La cabecera  correcta CERTUS")
        
        # Eliminar columnas F, G y H
        sheet.delete_cols(6, 3)
        formato_texto(sheet, 2, 3)
        formato_numero(sheet, 11, 12, 13, 14, 15, 16)
        formato_fecha(sheet, 7, 17, 18, 19, 20)

        for col_num, valor in enumerate(nueva_cabecera, 1):
            sheet.cell(row=1, column=col_num, value=valor)

        # Guardar los cambios en un nuevo archivo
        ruta_carpeta = define_carpeta('P CERTUS', fecha_actual)
        
        base_name = os.path.basename(filename)
        # Crear el nuevo nombre del archivo sin duplicar la ruta
        if 'CAMPUS' in base_name:
            new_base_name = f'PAGOS_CAMPUS_{fecha_actual}.xlsx'
        else:
            new_base_name = f'PAGOS_CERTUS_{fecha_actual}.xlsx'
        
        new_filename = os.path.join(ruta_carpeta, new_base_name)
        wb.save(new_filename)

        # Mover el archivo original a la carpeta especificada
        shutil.move(filename, ruta_carpeta)
        print("Certus Generado")
    else:
        print("La cabecera no es la esperada. Deteniendo la ejecución.")

def pagos_campus(filename, cabecera_esperada, nueva_cabecera,fecha_actual):
    wb_pre = load_workbook(filename)
    # sheet_pre = wb_pre.active

    if 1==1:
        print("La cabecera  correcta CERTUS")
        
        wb=reordenar_columnas(wb_pre,cab_campus_orden)
        sheet = wb.active
 
        formato_texto(sheet, 2, 3)
        formato_numero(sheet, 11, 12, 13, 14, 15, 16)
        formato_fecha(sheet, 7, 17, 18, 19, 20)
        
 
        for col_num, valor in enumerate(nueva_cabecera, 1):
            sheet.cell(row=1, column=col_num, value=valor)


        
        # Guardar los cambios en un nuevo archivo
        ruta_carpeta = define_carpeta('P CERTUS', fecha_actual)
        file_name = 'PAGOS_CAMPUS ' + fecha_actual + '.xlsx'
        new_filename = os.path.join(ruta_carpeta, file_name)
        wb.save(new_filename)

        # Abrir el archivo
        # os.startfile(new_filename)
        shutil.move(filename,ruta_carpeta)
        print("Certus Generado")
    else:
        print("La cabecera no es la esperada. Deteniendo la ejecución.")



# def pagos_tls(filename,fecha_actual):
#     wb = load_workbook(filename)
#     sheet = wb.active
    
#     formato_texto(sheet, 1)
#     formato_numero(sheet,buscar_posicion_por_nombre(sheet,'Importe Pagado'))
#     formato_fecha(sheet, buscar_posicion_por_nombre(sheet,'Fecha Pago'))
    
#     ruta_carpeta = define_carpeta('P TLS', fecha_actual)
#     file_name = 'PAGOS_TLS' + fecha_actual + '.xlsx'
#     new_filename = os.path.join(ruta_carpeta, file_name)
#     wb.save(new_filename)
    
#     # Abrir el archivo
#     ruta_v=r'V:\142. Ucal-Tls\\'
#     shutil.copy(new_filename,ruta_v)
#     # os.startfile(new_filename)
#     shutil.move(filename,ruta_carpeta)
#     print("TLS Generado")
    
    
    
def reordenar_columnas(archivo_excel, nuevo_orden):
    # Obtener la hoja pasada como argumento
    libro = archivo_excel
    hoja = libro.active

    # Insertar una columna 'NO_HAY_DATA' al principio sin datos
    hoja.insert_cols(1)
    hoja.cell(row=1, column=1, value='NO_HAY_DATA')

    # Obtener las posiciones actuales de las cabeceras
    cabeceras = [hoja.cell(row=1, column=col).value for col in range(1, hoja.max_column + 1)]
    posiciones_actuales = {cabecera: [] for cabecera in cabeceras}

    for index, cabecera in enumerate(cabeceras):
        posiciones_actuales[cabecera].append(index + 1)

    # Crear una nueva hoja para almacenar los datos reordenados
    nuevo_libro = Workbook()
    nueva_hoja = nuevo_libro.active

    # Crear las cabeceras en el nuevo orden
    for cabecera in nuevo_orden:
        nueva_hoja.cell(row=1, column=nuevo_orden.index(cabecera) + 1, value=cabecera)

    # Copiar los datos a la nueva hoja en el nuevo orden
    for row in hoja.iter_rows(min_row=2, max_row=hoja.max_row, values_only=True):
        nueva_fila = []
        for cabecera in nuevo_orden:
            # Obtener todas las posiciones de la cabecera en el archivo original
            posiciones = posiciones_actuales[cabecera]
            # Iterar sobre las posiciones y agregar el valor correspondiente
            for pos in posiciones:
                nueva_fila.append(row[pos - 1])
        nueva_hoja.append(nueva_fila)

    return nuevo_libro
     
    
    
# def reordenar_columnas(archivo_excel, nuevo_orden):
#     # Obtener la hoja pasada como argumento
#     libro = archivo_excel
#     hoja = libro.active

#     # Insertar una columna 'NO_HAY_DATA' al principio sin datos
#     hoja.insert_cols(1)
#     hoja.cell(row=1, column=1, value='NO_HAY_DATA')

#     # Obtener las posiciones actuales de las cabeceras
#     cabeceras = [hoja.cell(row=1, column=col).value for col in range(1, hoja.max_column + 1)]
#     posiciones_actuales = {cabecera: [] for cabecera in cabeceras}

#     for index, cabecera in enumerate(cabeceras):
#         posiciones_actuales[cabecera].append(index + 1)

#     # Crear una nueva hoja para almacenar los datos reordenados
#     nuevo_libro = Workbook()
#     nueva_hoja = nuevo_libro.active

#     # Crear las cabeceras en el nuevo orden
#     for cabecera in nuevo_orden:
#         nueva_hoja.cell(row=1, column=nuevo_orden.index(cabecera) + 1, value=cabecera)

#     # Copiar los datos a la nueva hoja en el nuevo orden
#     for row in hoja.iter_rows(min_row=2, max_row=hoja.max_row, values_only=True):
#         nueva_fila = []
#         for cabecera in nuevo_orden:
#             # Obtener todas las posiciones de la cabecera en el archivo original
#             posiciones = posiciones_actuales[cabecera]
#             # Iterar sobre las posiciones y agregar el valor correspondiente
#             for pos in posiciones:
#                 nueva_fila.append(row[pos - 1])
#         nueva_hoja.append(nueva_fila)

#     return nuevo_libro
 