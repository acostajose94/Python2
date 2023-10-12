import pandas as pd
import datetime
import os
import winreg
import shutil
from funciones_pagos import *
from formato_columnas import *

archivo1='Base Asignacion Inactiva II Mayo.xlsx'
archivo2='Base Asignacion Inactiva III Mayo.xlsx'
archivo5='Base Asignacion Inactiva IV Mayo.xlsx'
archivo3='Base Asignacion Diplomado, wetalk Mayo.xlsx'
archivo4='Base Asignacion Inactiva Formacion continua Mayo.xlsx'
#---------------------------

arch1=tabla(archivo1)

arch2=tabla(archivo2)
arch3=tabla(archivo3)
arch4=tabla(archivo4)
arch5=tabla(archivo5)
#tabla3=tabla('Base Asignacion Formacion continua Externa 2022  Enero.xlsx')

tabla_final = pd.concat([arch1,arch2,arch3,arch4,arch5])

# Obtener la fecha actual en formato dd_mm
fecha_actual = datetime.datetime.now().strftime('%d_%m')
# Crear la ruta de la carpeta en el escritorio con la fecha actual
# ruta_carpeta = os.path.join(os.path.expanduser('~'), 'Desktop',  fecha_actual+'p cibert')
# Este define la ruta si tienes onedrive, el anterior es si no lo tiene
ruta_carpeta = os.path.join(os.path.join(os.environ['USERPROFILE']), 'OneDrive', 'Escritorio',  fecha_actual+'p cibert Inactiva')
if not os.path.exists(ruta_carpeta):
    os.makedirs(ruta_carpeta)

ruta_archivo = os.path.join(ruta_carpeta, fecha_actual+' Pagos Inactiva.xlsx')
tabla_final.to_excel(ruta_archivo, index=False)

shutil.move(archivo1,ruta_carpeta)
shutil.move(archivo2,ruta_carpeta)
shutil.move(archivo3,ruta_carpeta)
shutil.move(archivo4,ruta_carpeta)
shutil.move(archivo5,ruta_carpeta)


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

workbook = load_workbook(ruta_archivo)
worksheet = workbook['Sheet1']

formato_fecha(worksheet, 7, 8, 34)
formato_texto(worksheet, 2, 3, 11, 13, 22, 23, 25, 27)
formato_numero(worksheet, 16,17,18,33)

# Recorre todas las columnas de la hoja de trabajo y ajusta el ancho
for columna in worksheet.columns:
    max_length = 0
    columna_letra = get_column_letter(columna[0].column)
    for celda in columna:
        if celda.value:
            if isinstance(celda.value, datetime):
                value_str = celda.value.strftime("%Y-%m-%d %H:%M:%S")  # Convertir a string con formato
                value_length = len(value_str)
                if value_length > max_length:
                    max_length = value_length
            else:
                value_length = len(str(celda.value))
                if value_length > max_length:
                    max_length = value_length
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[columna_letra].width = adjusted_width

# Guardar los cambios en el archivo de Excel
workbook.save(ruta_archivo)