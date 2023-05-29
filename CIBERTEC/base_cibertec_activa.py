import pandas as pd
import datetime
import os
import winreg
import shutil
from funciones_base import *
from formato_columnas import *
archivo1='Asignacion Cartera Activa 2023-1 Mayo 17-31 Mayo.xlsx'

#---------------------------

arch1=tabla_base(archivo1)
tabla_final = pd.concat([arch1])

# Obtener la fecha actual en formato dd_mm
fecha_actual = datetime.datetime.now().strftime('%d_%m')
ruta_carpeta = os.path.join(os.path.join(os.environ['USERPROFILE']), 'OneDrive', 'Escritorio',  fecha_actual+'b cibert Activa')
if not os.path.exists(ruta_carpeta):
    os.makedirs(ruta_carpeta)

ruta_archivo = os.path.join(ruta_carpeta, fecha_actual+' Base Activa.xlsx')
tabla_final.to_excel(ruta_archivo, index=False)
shutil.move(archivo1,ruta_carpeta)


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# Load the Excel file
workbook = load_workbook(ruta_archivo)
worksheet = workbook['Sheet1']

formato_fecha(worksheet, 7, 8)
formato_texto(worksheet, 2, 3, 11, 13, 22, 23, 25, 27)
formato_numero(worksheet, 16,17,18)


# Recorre todas las columnas de la hoja de trabajo y ajusta el ancho
for columna in worksheet.columns:
    max_length = 0
    columna_letra = get_column_letter(columna[0].column)
    for celda in columna:
        if celda.value:
            if isinstance(celda.value, datetime):
                value_str = celda.value.strftime("%Y-%m-%d %H:%M:%S")  # Convertir a string con formato
                value_length = len(value_str)
                if value_length > max_length:
                    max_length = value_length
            else:
                value_length = len(str(celda.value))
                if value_length > max_length:
                    max_length = value_length
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[columna_letra].width = adjusted_width


# Guardar los cambios en el archivo de Excel
workbook.save(ruta_archivo)
