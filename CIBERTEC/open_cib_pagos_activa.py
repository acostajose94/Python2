import openpyxl

# Load the Excel file
workbook = openpyxl.load_workbook('Asignacion Cartera Activa 2023-1 Mayo 17-31 Mayo.xlsx')

# Select the worksheet
worksheet = workbook['Sheet1']

# Remove the first row (header)
worksheet.delete_rows(1, 1)

# Get the header row
header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

# Get the starting position of the 'IMPORTE' and 'FECHA' columns
importe_start = header_row.index('IMPORTE')
fecha_start = header_row.index('FECHA')

# Create a new dictionary to store the grouped data
grouped_data = {}

# Loop through each row in the worksheet
for row in worksheet.iter_rows(min_row=2, values_only=True):
    importe = row[importe_start]
    fecha = row[fecha_start]

    # Create a key based on 'IMPORTE' and 'FECHA'
    key = (importe, fecha)

    # Check if the key already exists in the dictionary
    if key in grouped_data:
        grouped_data[key].append(row)
    else:
        grouped_data[key] = [row]

# Guardar el nuevo libro de trabajo como "datos_agrupados.xlsx"
workbook.save('datos_agrupados.xlsx')