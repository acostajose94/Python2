
from openpyxl.styles import numbers
from openpyxl import load_workbook
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
def formato_texto(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]
    
    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            cell = row[0]
            cell.number_format = '@'
            if isinstance(cell.value, (int, float)):
                cell.value = str(cell.value)
            
def formato_numero(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]
        
    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            cell = row[0]
            valor = cell.value
            if valor is not None and isinstance(valor, str):
                valor = valor.replace(',', '.')
                cell.value = float(valor)
            cell.number_format = '0.00'

def formato_fecha(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]

    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            for cell in row:
                # Verificar si el valor termina en "/202"
                valor = str(cell.value)
                if valor.endswith("/202"):
                    nuevo_valor = valor[:-4] + "2"
                    cell.value = nuevo_valor

                # Aplicar formato de fecha
                cell.number_format = numbers.FORMAT_DATE_XLSX14

def ajusta_fecha(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]

    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            for cell in row:
                # Verificar si el valor termina en "/202"
                valor = str(cell.value)
                if valor.endswith("/202"):
                    nuevo_valor = valor[:-4] + "2"
                    cell.value = nuevo_valor

                # Aplicar formato de fecha
                cell.number_format = numbers.FORMAT_DATE_XLSX14

 


def formato_columnas(ruta_archivo):
    # Load the Excel file
    workbook = load_workbook(ruta_archivo)
    worksheet = workbook['Sheet1']

    ajusta_fecha(worksheet,34)
    formato_fecha(worksheet, 7, 8, 34)
    formato_texto(worksheet, 2, 3, 11, 13, 22, 23, 25, 27)
    formato_numero(worksheet, 16, 17, 18, 33)

    # Recorre todas las columnas de la hoja de trabajo y ajusta el ancho
    for columna in worksheet.columns:
        max_length = 0
        columna_letra = get_column_letter(columna[0].column)
        for celda in columna:
            if celda.value:
                if isinstance(celda.value, datetime):
                    value_str = celda.value.strftime("%Y-%m-%d %H:%M:%S")  # Convertir a string con formato
                    value_length = len(value_str)
                    if value_length > max_length:
                        max_length = value_length
                else:
                    value_length = len(str(celda.value))
                    if value_length > max_length:
                        max_length = value_length
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[columna_letra].width = adjusted_width

    # Guardar los cambios en el archivo de Excel
    workbook.save(ruta_archivo)
    
    