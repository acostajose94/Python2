import openpyxl

def abrir_excel():
    try:
        # Intentar abrir el archivo de Excel
        archivo = openpyxl.load_workbook('archivo.xlsx')
    except FileNotFoundError:
        return None

    # Buscar las hojas en el orden especificado
    for nombre_hoja in ['Base de datos', 'Base de datos ', 'DB']:
        if nombre_hoja in archivo.sheetnames:
            return archivo[nombre_hoja]

    # Si no se encontró ninguna hoja válida, devolver None
    return None


def procesar_datos():
    if 'archivo.xlsx' in os.listdir():
        hoja = abrir_excel()
    else:
        return None

    if hoja is None:
        return None

    cabecera = [celda.value for celda in hoja[1]]

    columna_gestor = None
    for idx, valor in enumerate(cabecera):
        if valor == 'Gestor':
            columna_gestor = idx + 1
            break

    if columna_gestor is None:
        return None

    for fila in hoja.iter_rows(min_row=2):
        valor_gestor = fila[columna_gestor - 1].value
        if valor_gestor is not None:
            fila[columna_gestor - 1].value = valor_gestor.lower()

            if 'aval' in valor_gestor.lower():
                valor_aval = fila[columna_gestor - 1].value
                fila[columna_gestor - 1].value = valor_aval.capitalize()

    hoja.parent.save('archivo.xlsx')

procesar_datos()
