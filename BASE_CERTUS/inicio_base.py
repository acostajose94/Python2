import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from funciones import *
from variables import *

archivo_reciente=extraer_archivo_mas_reciente()

retira=obtener_excel('Retirar')
if retira is not None:
    shutil.copy(retira, destino_retirar)
    os.remove(retira)
    
# repacta=obtener_excel('REPAC')
# if repacta is not None:
#     shutil.copy(repacta, destino_retirar)
#     os.remove(repacta)

archivos_excel=listar_excels()
ruta_carpeta=define_carpeta('B CERTUS',fecha_actual)

print (ruta_carpeta)
contador=0
for archivo_excel in archivos_excel:
    
    
    
    # Abre el archivo de Excel
    workbook = openpyxl.load_workbook(archivo_excel)
    sheet = workbook.active
    
    
    if 'CAMPUS' in archivo_excel:
        contador=contador+1
        #-------EXCEL PARA EXTRA---------#
        #Nuebo book para EXTRA
        nuevo_wb2 =   load_workbook(filename=archivo_excel)
        
        nuevo_sheet2 = nuevo_wb2.active
        
        nuevo_sheet2.title = "Hoja1"
        # Reemplazar espacio en cabecera
        for celda in nuevo_sheet2[1]:
            if isinstance(celda.value, str):  # Verifica si el valor de la celda es una cadena
                # Elimina los espacios en el valor de la celda y sobrescribe la celda
                celda.value = celda.value.replace(" ", "")
        
        nuevo_wb2=reordenar_columnas(nuevo_wb2,cabecera_extra)
        
        nuevo_sheet2 = nuevo_wb2.active
        nuevo_sheet2.title = "Hoja1"
        formato_texto(nuevo_sheet2,1,2,3,4,5,6)
        ruta_base_com2=os.path.join(ruta_carpeta, 'extra'+str(contador))
        nuevo_wb2.save(ruta_base_com2 + '.xlsx')
        
        
        
        # Reemplazar espacio en cabecera
        for celda in sheet[1]:
            if isinstance(celda.value, str):  # Verifica si el valor de la celda es una cadena
                # Elimina los espacios en el valor de la celda y sobrescribe la celda
                celda.value = celda.value.replace(" ", "")
            
        ver_cab=verificar_cabecera_completa(sheet, cabecera_esperada_campus)    
        print(ver_cab)
        if ver_cab is False:
            break
        #reordenar columnas
        workbook=reordenar_columnas(workbook, cabecera_orden_campus)
        sheet = workbook.active
        
        
        #reemplazar cabecera
        for col_idx, valor in enumerate(cabecera_esperada, start=1):
            sheet.cell(row=1, column=col_idx, value=valor)
        
        # Asignar los valores  
        total_filas = sheet.max_row
        
        contenido_columna = 'DNI'
        for fila in range(2, total_filas + 1):  # Comenzar desde la segunda fila hasta la última fila
            sheet.cell(row=fila, column=3, value=contenido_columna)
        
        contenido_columna = 'Y'
        for fila in range(2, total_filas + 1):  # Comenzar desde la segunda fila hasta la última fila
            sheet.cell(row=fila, column=45, value=contenido_columna)
            
        contenido_columna = 0
        for fila in range(2, total_filas + 1):  # Comenzar desde la segunda fila hasta la última fila
            sheet.cell(row=fila, column=44, value=contenido_columna)    
        
        
        contenido_columna = 'NUEVO'
        for fila in range(2, total_filas + 1):  # Comenzar desde la segunda fila hasta la última fila
            sheet.cell(row=fila, column=17, value=contenido_columna) 
    
        contenido_columna = 'C'
        for fila in range(2, total_filas + 1):  # Comenzar desde la segunda fila hasta la última fila
            sheet.cell(row=fila, column=18, value=contenido_columna) 
    
    
    verificar_cabecera_completa(sheet, cabecera_esperada)
    sheet=insertar_columnas_vacias(sheet, nombre_columna, cantidad_columnas)
     
    # elimina_tabulaciones(sheet)
    
    elimina_tabulaciones_columnas(sheet, 12)   
    formato_texto(sheet, 1,2, 4, 7, 8, 9,10,16,35)
    formato_numero(sheet, 36,37,38,39,40,41,42,43,44)
    formato_fecha(sheet,25,27,31,32)
    # eliminar_columnas(sheet, 47, 48)
    # Después de formato_fecha(sheet, 25, 27, 31, 32)
    
    columnas_a_eliminar = ['COD. DUPL',	'EXT HBC']
    eliminar_columnas_por_nombre(sheet, columnas_a_eliminar)

        
    # Asignar los valores de la nueva cabecera
    for col_idx, valor in enumerate(nueva_cabecera, start=1):
        sheet.cell(row=1, column=col_idx, value=valor)

    # posicion_cabecera = 49
    # contenido_cabecera = 'AVAL 2023 PREGRADO'
    
    # # Obtén la cantidad de filas en la hoja
    total_filas = sheet.max_row

    # # Agregar 'AVAL 2023 PREGRADO' en todas las filas de la columna 49
    contenido_columna = 'AVAL'

    for fila in range(2, total_filas + 1):  # Comenzar desde la segunda fila hasta la última fila
        sheet.cell(row=fila, column=49, value=contenido_columna)
    
    corregir_valor_columna(sheet, 22, 'MA?ANA', 'MAÑANA')
    
    columna_letra = 'L'
    # Recorrer la columna y eliminar los caracteres especificados
    for celda in sheet[columna_letra]:
        # Reemplazar los caracteres por una cadena vacía
        celda.value = str(celda.value).replace(chr(10), '').replace(chr(9), '').replace(chr(13), '')


    
    nombre_short=obtener_despues_de_pregrado(archivo_excel)
    ruta_base=os.path.join(ruta_carpeta, 'BASE CERTUS ' + archivo_excel)
    workbook.save(ruta_base + '.xlsx')
    shutil.move(archivo_excel,ruta_carpeta)
    # guardar_archivo_mover('B CERTUS', workbook, fecha_actual,nombre_short,archivo_excel,retira)
    print(nombre_short)
# shutil.move(retira,ruta_carpeta)
# os.remove(archivo_reciente)