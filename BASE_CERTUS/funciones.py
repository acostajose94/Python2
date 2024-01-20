
from openpyxl.styles import numbers
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
import os
import shutil


import glob
import zipfile
import rarfile

def extraer_archivo_mas_reciente():
    # Obtener la ruta del archivo Python actual
    ruta_actual = os.path.dirname(os.path.abspath(__file__))

    # Buscar archivos ZIP y RAR en la carpeta actual
    archivos_zip = glob.glob(os.path.join(ruta_actual, '*.zip'))
    archivos_rar = glob.glob(os.path.join(ruta_actual, '*.rar'))

    # Combinar y ordenar los archivos por fecha de modificación descendente
    archivos = sorted(archivos_zip + archivos_rar, key=os.path.getmtime, reverse=True)

    if archivos:
        archivo_mas_reciente = archivos[0]
        nombre_archivo, extension = os.path.splitext(archivo_mas_reciente)

        # Extraer archivo ZIP
        if extension == '.zip':
            with zipfile.ZipFile(archivo_mas_reciente, 'r') as archivo_zip:
                archivo_zip.extractall(ruta_actual)
                print(f'Se ha extraído el archivo ZIP más reciente: {archivo_mas_reciente}')

        # Extraer archivo RAR
        elif extension == '.rar':
            with rarfile.RarFile(archivo_mas_reciente, 'r') as archivo_rar:
                archivo_rar.extractall(ruta_actual)
                print(f'Se ha extraído el archivo RAR más reciente: {archivo_mas_reciente}')
        return archivo_mas_reciente

    else:
        print('No se encontraron archivos ZIP o RAR en la carpeta actual.')
   
def listar_excels():
    archivos_excel = []
    directorio_actual = os.getcwd()

    for archivo in os.listdir(directorio_actual):
        if archivo.endswith('.xlsx') or archivo.endswith('.xls'):
            if 'Retira' not in archivo:
                archivos_excel.append(archivo)
    
    return archivos_excel        
 
def obtener_excel_retira():
    directorio_actual = os.getcwd()

    for archivo in os.listdir(directorio_actual):
        if archivo.endswith('.xlsx') or archivo.endswith('.xls'):
            if 'Retira' in archivo:
                return archivo
    
    return None

def obtener_despues_de_pregrado(archivo):
    palabra_clave = 'CERTUS_'
    indice = archivo.lower().find(palabra_clave.lower())

    if indice != -1:
        indice += len(palabra_clave)
        return archivo[indice:].strip()
    
    return None   
 
def verificar_cabecera_completa(sheet, cabecera):
    # Obtener la primera fila de la hoja de trabajo (cabecera)
    primera_fila = sheet[1]
    
    # Verificar si la cabecera completa existe
    for celda, valor_cabecera in zip(primera_fila, cabecera):
        if celda.value != valor_cabecera:
            print(f"La columna {celda.column_letter} no coincide: se esperaba '{valor_cabecera}', se encontró '{celda.value}'")
            return False
            break
    
    return True

def insertar_columnas_vacias(sheet, nombre_columna, cantidad_columnas):
    # Encuentra el índice de la columna especificada
    columna_referencia = None
    for columna in sheet.iter_cols():
        if columna[0].value == nombre_columna:
            columna_referencia = columna[0].column
            break

    if columna_referencia is not None:
        # Inserta las columnas vacías después de la columna de referencia
        sheet.insert_cols(columna_referencia + 1, amount=cantidad_columnas)

    return sheet

def define_carpeta(nombre_carpeta,fecha_actual):
    ruta_carpeta = os.path.join(os.path.join(os.environ['USERPROFILE']), 'OneDrive', 'Escritorio', fecha_actual + nombre_carpeta)
    if not os.path.exists(ruta_carpeta):
        os.makedirs(ruta_carpeta)
    return ruta_carpeta


def guardar_archivo_mover(nombre_carpeta, workbook, fecha_actual, nombre_a_guardar, *archivos_mover):
    ruta_carpeta = os.path.join(os.path.join(os.environ['USERPROFILE']), 'OneDrive', 'Escritorio', fecha_actual + nombre_carpeta)
    if not os.path.exists(ruta_carpeta):
        os.makedirs(ruta_carpeta)

    ruta_archivo = os.path.join(ruta_carpeta, 'BASE' + nombre_a_guardar)
    workbook.save(ruta_archivo)

    for arch in archivos_mover:
        try:
            shutil.move(arch, ruta_carpeta)
            print(f"El archivo {arch} se ha movido exitosamente a {ruta_carpeta}")
        except FileNotFoundError:
            print(f"No se pudo encontrar el archivo {arch}. Se omitirá.")

# Desde Aqui 

def elimina_tabulaciones(sheet: Worksheet):
    # Buscar la columna por su nombre
    columna_direccion = None
    for columna in sheet.iter_cols(values_only=True):
        if columna[0] == 'DIRECCION':
            columna_direccion = columna
            break

    # Reemplazar la tabulación de salto de línea con un espacio en la columna 'DIRECCION'
    if columna_direccion is not None:
        for cell in columna_direccion[1:]:  # Comenzar desde la segunda fila, asumiendo que la primera fila son encabezados
            if isinstance(cell, Cell) and cell.value is not None and '\t\n' in cell.value:
                cell.value = cell.value.replace('\t\n', ' ')


def formato_texto(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]
    
    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            cell = row[0]
            cell.number_format = '@'
            if isinstance(cell.value, (int, float)):
                cell.value = str(cell.value)
            
def formato_numero(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]
        
    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            cell = row[0]
            valor = cell.value
            if valor is not None and isinstance(valor, str):
                valor = valor.replace(',', '.')
                cell.value = float(valor)
            cell.number_format = '0.00'

def formato_fecha(sheet, *columnas):
    if len(columnas) == 1 and isinstance(columnas[0], list):
        columnas = columnas[0]

    for columna in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=columna, max_col=columna):
            for cell in row:
                # Verificar si el valor termina en "/202"
                valor = str(cell.value)
                if valor.endswith("/202"):
                    nuevo_valor = valor[:-4] + "2"
                    cell.value = nuevo_valor

                # Aplicar formato de fecha
                cell.number_format = numbers.FORMAT_DATE_XLSX14


# Define una función para eliminar columnas por nombre
def eliminar_columnas_por_nombre(sheet, columnas_a_eliminar):
    for col_name in columnas_a_eliminar:
        for col in sheet.columns:
            if col[0].value == col_name:
                sheet.delete_cols(col[0].column)


def elimina_tabulaciones_columnas(sheet: Worksheet, *columnas):
    for col_idx in columnas:
        for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell,Cell) and cell.value is not None and '\t\n' in cell.value:
                    cell.value = cell.value.replace('\t\n', ' ')              
             
def corregir_valor_columna(sheet, posicion_columna, valor_a_reemplazar, nuevo_valor):
    # Obtén la cantidad de filas en la hoja
    total_filas = sheet.max_row

    for fila in range(2, total_filas + 1):
        valor_celda = sheet.cell(row=fila, column=posicion_columna).value
        if valor_celda and valor_a_reemplazar in valor_celda:
            nuevo_valor_celda = valor_celda.replace(valor_a_reemplazar, nuevo_valor)
            sheet.cell(row=fila, column=posicion_columna, value=nuevo_valor_celda)

