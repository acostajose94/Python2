import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from b_funciones_ALF import *
from b_variables_ALF import *

# archivo_reciente=extraer_archivo_mas_reciente()
archivos_excel=listar_excels()

ruta_carpeta=define_carpeta('AFL B',fecha_actual)

for archivo_excel in archivos_excel:

    pre_workbook=load_workbook(filename=archivo_excel)
    sheet_PRE = pre_workbook.active
    ver_cab=verificar_cabecera_completa(sheet_PRE, cab_esperada)
    print(ver_cab)
    if ver_cab is False:
        print("Stoped")
        break
    workbook=reordenar_columnas(pre_workbook,cab_ordenada)
    sheet = workbook.active
    
    formato_texto(sheet,2,4,6,20,21,22,23,24)
    # formato_numero(sheet, 11,12,13,14,16,17,18,19,20)
    # formato_fecha(sheet, 8,9,10,21,24)

    # Asignar los valores de la nueva cabecera
    for col_idx, valor in enumerate(cab_nueva, start=1):
        sheet.cell(row=1, column=col_idx, value=valor)

    # columna_letra = 'AD'
    # # Recorrer la columna y eliminar los caracteres especificados
    # for celda in sheet[columna_letra]:
    #     # Reemplazar los caracteres por una cadena vacía
    #     celda.value = str(celda.value).replace(chr(10), '').replace(chr(9), '').replace(chr(13), '')


    ruta_base=os.path.join(ruta_carpeta, archivo_excel)
    workbook.save(ruta_base + '.xlsx')
    shutil.move(archivo_excel,ruta_carpeta)

    print(ruta_base)
 
print('Echo')
 