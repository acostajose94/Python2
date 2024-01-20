import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def listar_archivos_txt():
    archivos_txt = []
    directorio_actual = os.getcwd()

    for archivo in os.listdir(directorio_actual):
        if archivo.endswith('.txt'):
            archivos_txt.append(archivo)
    return archivos_txt

# Directorio donde se encuentran los archivos TXT
directorio = listar_archivos_txt()

# Crear un nuevo archivo de Excel
libro_excel = Workbook()
hoja = libro_excel.active
hoja.title = 'Datos'

# Iterar sobre los archivos TXT en el directorio
for archivo in directorio:
    if archivo.endswith('.txt'):
        ruta_archivo = os.path.join(directorio, archivo)
        with open(ruta_archivo, 'r') as archivo_txt:
            lineas = archivo_txt.readlines()
            # Escribir los datos en el archivo de Excel
            columna = os.path.splitext(archivo)[0]  # Nombre del archivo sin la extensión
            columna_letra = get_column_letter(len(hoja[1]) + 1)  # Obtener la letra de la columna
            hoja[columna_letra + '1'] = columna  # Escribir el nombre del archivo en la primera fila
            for i, linea in enumerate(lineas, start=1):
                hoja[columna_letra + str(i + 1)] = linea.strip()  # Escribir cada línea en la columna correspondiente

# Guardar el archivo de Excel
libro_excel.save('datos_completos.xlsx')
print("¡Archivos TXT unificados y guardados en un archivo Excel!")
